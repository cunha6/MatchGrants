"""Testes da app planned_grants. A rede (página do plano + download do Excel) é sempre MOCKADA —
a suite corre offline. O Excel é simulado com um Workbook openpyxl em memória.
"""

import os
from datetime import date, datetime, timedelta
from decimal import Decimal
from unittest import mock

from django.contrib.auth.models import User
from django.test import TestCase
from openpyxl import Workbook

from . import services
from .models import PlannedGrant

TEST_PASSWORD = os.environ.get("TEST_USER_PASSWORD", "test-only-password")

_HEADER = [
    "ID", "Tipo Ent. Beneficiária", "Natureza Aviso", "Designação do Aviso", "Programa",
    "Objetivo Específico", "Fundo", "Dotação Fundo", "Data Início Prevista", "Data Fim Prevista",
    "Quadrimestre", "NUTS II", "Modalidade Apresentação Candidatura",
]


def _workbook(rows):
    """Workbook em memória com o cabeçalho (linha 1, ignorada) + as `rows` dadas."""
    wb = Workbook()
    ws = wb.active
    ws.append(_HEADER)
    for row in rows:
        ws.append(list(row))
    return wb


def _row(plan_id, designation="Aviso X", allocation="990 000 €",
         start=None, end=None, programme="PT2030", grant_type="Concurso"):
    """Uma linha do Excel na ordem correta das colunas."""
    return [plan_id, "PME", grant_type, designation, programme, "OE 1.1", "FEDER",
            allocation, start, end, "1.º Quadrimestre", "Norte", "Balcão dos Fundos"]


# --- Normalização --------------------------------------------------------

class NormalizationTests(TestCase):
    def test_decimal_portuguese_formats(self):
        self.assertEqual(services._parse_decimal("990 000 €"), Decimal("990000"))
        self.assertEqual(services._parse_decimal("1.234.567,89 €"), Decimal("1234567.89"))
        self.assertEqual(services._parse_decimal("990000,50"), Decimal("990000.50"))
        self.assertEqual(services._parse_decimal(990000.5), Decimal("990000.5"))

    def test_decimal_empty_and_garbage(self):
        self.assertIsNone(services._parse_decimal(""))
        self.assertIsNone(services._parse_decimal(None))
        self.assertIsNone(services._parse_decimal("sem valor"))

    def test_date_from_excel_datetime_and_text(self):
        self.assertEqual(services._parse_date(datetime(2026, 4, 30, 12, 0)), date(2026, 4, 30))
        self.assertEqual(services._parse_date(date(2026, 4, 30)), date(2026, 4, 30))
        self.assertEqual(services._parse_date("30/04/2026"), date(2026, 4, 30))
        self.assertIsNone(services._parse_date(""))
        self.assertIsNone(services._parse_date(None))

    def test_int_tolerant(self):
        self.assertEqual(services._parse_int(12), 12)
        self.assertEqual(services._parse_int("12"), 12)
        self.assertEqual(services._parse_int("12.0"), 12)
        self.assertIsNone(services._parse_int(""))
        self.assertIsNone(services._parse_int("abc"))

    def test_is_invitation(self):
        self.assertTrue(services._is_invitation("Convite"))
        self.assertTrue(services._is_invitation("convite "))
        self.assertFalse(services._is_invitation("Concurso"))
        self.assertFalse(services._is_invitation(""))
        self.assertFalse(services._is_invitation(None))


# --- Descoberta do Excel -------------------------------------------------

class FindXlsxTests(TestCase):
    def _resp(self, html):
        m = mock.Mock()
        m.raise_for_status.return_value = None
        m.text = html
        return m

    def test_finds_first_xlsx_and_resolves_relative(self):
        html = ('<a href="/docs/guia.pdf">guia</a>'
                '<a href="/ficheiros/plano-2026.xlsx">plano</a>'
                '<a href="/outro.xlsx">outro</a>')
        with mock.patch("planned_grants.services.requests.get", return_value=self._resp(html)):
            url = services._find_xlsx_url()
        self.assertEqual(url, "https://portugal2030.pt/ficheiros/plano-2026.xlsx")

    def test_raises_when_no_xlsx(self):
        with mock.patch("planned_grants.services.requests.get",
                        return_value=self._resp('<a href="/x.pdf">x</a>')):
            with self.assertRaises(services.PlannedGrantsSyncError):
                services._find_xlsx_url()


# --- Sincronização -------------------------------------------------------

class SyncTests(TestCase):
    def _sync(self, rows):
        with mock.patch("planned_grants.services._download_workbook",
                        return_value=_workbook(rows)):
            services.sync_planned_grants()

    def test_creates_records_and_maps_fields(self):
        self._sync([_row(1, designation="Inovação Produtiva",
                          allocation="990 000 €", start=datetime(2026, 5, 1), end=datetime(2026, 9, 30))])
        pg = PlannedGrant.objects.get(plan_id=1)
        self.assertEqual(pg.designation, "Inovação Produtiva")
        self.assertEqual(pg.beneficiary_type, "PME")
        self.assertEqual(pg.total_allocation, Decimal("990000"))
        self.assertEqual(pg.expected_start, date(2026, 5, 1))
        self.assertEqual(pg.expected_end, date(2026, 9, 30))
        self.assertEqual(pg.nuts, "Norte")

    def test_row_without_numeric_id_is_skipped(self):
        self._sync([_row(1), _row("", designation="linha lixo"), [None] * 13])
        self.assertEqual(PlannedGrant.objects.count(), 1)

    def test_idempotent_when_unchanged(self):
        self._sync([_row(1)])
        before = PlannedGrant.objects.get(plan_id=1).updated_at
        self._sync([_row(1)])                       # exatamente igual
        after = PlannedGrant.objects.get(plan_id=1).updated_at
        self.assertEqual(before, after)             # nada foi escrito
        self.assertEqual(PlannedGrant.objects.count(), 1)

    def test_updates_only_when_a_field_changes(self):
        self._sync([_row(1, designation="Original")])
        before = PlannedGrant.objects.get(plan_id=1).updated_at
        self._sync([_row(1, designation="Alterado")])
        pg = PlannedGrant.objects.get(plan_id=1)
        self.assertEqual(pg.designation, "Alterado")
        self.assertGreaterEqual(pg.updated_at, before)

    def test_second_sync_adds_new_and_keeps_old(self):
        self._sync([_row(1)])
        self._sync([_row(1), _row(2)])
        self.assertEqual(set(PlannedGrant.objects.values_list("plan_id", flat=True)), {1, 2})

    def test_invitations_are_not_added(self):
        self._sync([_row(1, grant_type="Concurso"), _row(2, grant_type="Convite")])
        self.assertEqual(set(PlannedGrant.objects.values_list("plan_id", flat=True)), {1})

    def test_existing_row_that_becomes_invitation_is_removed(self):
        self._sync([_row(1, grant_type="Concurso")])
        self.assertTrue(PlannedGrant.objects.filter(plan_id=1).exists())
        self._sync([_row(1, grant_type="Convite")])   # passou a convite → removido
        self.assertFalse(PlannedGrant.objects.filter(plan_id=1).exists())


# --- Serializer ----------------------------------------------------------

class SerializerTests(TestCase):
    def test_has_required_fields(self):
        pg = PlannedGrant.objects.create(
            plan_id=7, designation="D", programme="P", fund="FEDER",
            total_allocation=Decimal("1000.00"), nuts="Centro",
            expected_start=date(2026, 6, 1), expected_end=date(2026, 8, 1),
        )
        data = services.serialize_planned_grant(pg)
        self.assertEqual(
            set(data),
            {"id", "plan_id", "designation", "programme", "expected_start",
             "expected_end", "fund", "budget", "nuts"},
        )
        self.assertEqual(data["budget"], 1000.0)
        self.assertEqual(data["expected_start"], "2026-06-01")


# --- Endpoints -----------------------------------------------------------

class EndpointTests(TestCase):
    def setUp(self):
        # A listagem (Plano Anual) exige sessão admin/commercial; o sync fica aberto (automação).
        user = User.objects.create_user("comercial_plano", password=TEST_PASSWORD)
        user.profile.role = "commercial_grants"
        user.profile.save()
        self.client.force_login(user)

    def test_list_requires_authentication(self):
        self.client.logout()
        self.assertEqual(self.client.get("/planned-grants/").status_code, 401)

    def test_list_rejects_client_role(self):
        self.client.logout()
        client_user = User.objects.create_user("cliente_plano", password=TEST_PASSWORD)
        self.client.force_login(client_user)  # role=client por omissão (signal)
        self.assertEqual(self.client.get("/planned-grants/").status_code, 403)

    def test_sync_works_without_authentication(self):
        self.client.logout()
        with mock.patch("planned_grants.services._download_workbook",
                        return_value=_workbook([_row(1)])):
            resp = self.client.get("/planned-grants/sync/")
        self.assertEqual(resp.status_code, 200)

    def test_sync_returns_success(self):
        with mock.patch("planned_grants.services._download_workbook",
                        return_value=_workbook([_row(1)])):
            resp = self.client.get("/planned-grants/sync/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json(), {"success": True})
        self.assertEqual(PlannedGrant.objects.count(), 1)

    def test_sync_reports_scrape_failure(self):
        with mock.patch("planned_grants.services._download_workbook",
                        side_effect=services.PlannedGrantsSyncError("sem xlsx")):
            resp = self.client.get("/planned-grants/sync/")
        self.assertEqual(resp.status_code, 502)
        self.assertFalse(resp.json()["success"])

    def test_sync_rejects_post(self):
        self.assertEqual(self.client.post("/planned-grants/sync/").status_code, 405)

    def test_list_returns_paginated_envelope(self):
        today = date.today()
        for i in range(3):
            PlannedGrant.objects.create(plan_id=100 + i, designation=f"Aviso {i}",
                                        expected_start=today + timedelta(days=i + 1))
        resp = self.client.get("/planned-grants/?page_size=2")
        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertEqual(body["total"], 3)
        self.assertEqual(body["page_size"], 2)
        self.assertEqual(body["num_pages"], 2)
        self.assertEqual(len(body["planned_grants"]), 2)
        # Ordenado por data de abertura prevista (Meta.ordering).
        self.assertEqual([p["plan_id"] for p in body["planned_grants"]], [100, 101])

    def test_list_hides_past_expected_start(self):
        today = date.today()
        PlannedGrant.objects.create(plan_id=200, designation="Passado",
                                    expected_start=today - timedelta(days=1))
        PlannedGrant.objects.create(plan_id=201, designation="Hoje", expected_start=today)
        PlannedGrant.objects.create(plan_id=202, designation="Futuro",
                                    expected_start=today + timedelta(days=30))
        PlannedGrant.objects.create(plan_id=203, designation="Sem data", expected_start=None)
        body = self.client.get("/planned-grants/").json()
        self.assertEqual({p["plan_id"] for p in body["planned_grants"]}, {201, 202})

    def test_list_rejects_post(self):
        self.assertEqual(self.client.post("/planned-grants/").status_code, 405)

    def _for_ordering(self):
        today = date.today()
        PlannedGrant.objects.create(
            plan_id=300, designation="Cedo/Curto/PoucaVerba",
            expected_start=today + timedelta(days=5), expected_end=today + timedelta(days=10),
            total_allocation=Decimal("1000.00"))
        PlannedGrant.objects.create(
            plan_id=301, designation="Tarde/Longo/MuitaVerba",
            expected_start=today + timedelta(days=60), expected_end=today + timedelta(days=200),
            total_allocation=Decimal("999999.00"))

    def _order(self, order_by):
        body = self.client.get(f"/planned-grants/?order_by={order_by}").json()
        return [p["plan_id"] for p in body["planned_grants"]]

    def test_default_ordering_is_start_earliest(self):
        self._for_ordering()
        self.assertEqual(self._order(""), [300, 301])

    def test_ordering_start_latest(self):
        self._for_ordering()
        self.assertEqual(self._order("start_latest"), [301, 300])

    def test_ordering_end_earliest_and_latest(self):
        self._for_ordering()
        self.assertEqual(self._order("end_earliest"), [300, 301])
        self.assertEqual(self._order("end_latest"), [301, 300])

    def test_ordering_allocation_highest_and_lowest(self):
        self._for_ordering()
        self.assertEqual(self._order("allocation_highest"), [301, 300])
        self.assertEqual(self._order("allocation_lowest"), [300, 301])

    def test_unknown_order_by_falls_back_to_default(self):
        self._for_ordering()
        self.assertEqual(self._order("nonsense"), [300, 301])
