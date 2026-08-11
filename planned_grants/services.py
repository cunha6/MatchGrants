"""Sincronização do Plano Anual de Avisos (Portugal 2030) e newsletter semanal.

Fluxo do sync (`sync_planned_grants`):
  1. GET à página do plano anual;
  2. encontra automaticamente o 1.º link `<a>` cujo href termina em `.xlsx` (sem URL fixa);
  3. descarrega o Excel para MEMÓRIA (sem ficheiros temporários);
  4. lê com openpyxl (ignorando a linha de cabeçalho);
  5. faz upsert por `plan_id`, atualizando SÓ os campos que mudaram.

O serializer `serialize_planned_grant` também vive aqui (é a serialização de um `PlannedGrant`); a
newsletter semanal, que agrega várias entidades, está na app `newsletter`.
"""

import io
import logging
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from django.db import transaction
from openpyxl import load_workbook

from common.dates import parse_date
from .models import PlannedGrant

logger = logging.getLogger(__name__)

PLAN_URL = "https://portugal2030.pt/plano-anual-de-avisos/"
_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}
_TIMEOUT = 30  # segundos

# Ordenações suportadas pela listagem (?order_by=). Default (Meta.ordering): abertura mais cedo.
ORDERING = {
    "start_earliest":  "expected_start",
    "start_latest":    "-expected_start",
    "end_earliest":    "expected_end",
    "end_latest":      "-expected_end",
    "allocation_highest": "-total_allocation",
    "allocation_lowest":  "total_allocation",
}

# Ordem das colunas do Excel → campo do modelo. O índice é a posição da coluna (0-based); a
# coluna 0 (ID) é tratada à parte (é a chave do upsert). Ver o cabeçalho do plano anual.
_COLUMN_ORDER = [
    "beneficiary_type",     # 1  Tipo Ent. Beneficiária
    "grant_type",           # 2  Natureza Aviso
    "designation",          # 3  Designação do Aviso
    "programme",            # 4  Programa
    "specific_objective",   # 5  Objetivo Específico
    "fund",                 # 6  Fundo
    "total_allocation",     # 7  Dotação Fundo
    "expected_start",       # 8  Data Início Prevista
    "expected_end",         # 9  Data Fim Prevista
    "quadrimester",         # 10 Quadrimestre
    "nuts",                 # 11 NUTS II
    "submission_mode",      # 12 Modalidade Apresentação Candidatura
]
# Campos que precisam de conversão de tipo (os restantes são texto limpo).
_DECIMAL_FIELDS = {"total_allocation"}
_DATE_FIELDS = {"expected_start", "expected_end"}


class PlannedGrantsSyncError(Exception):
    """Falha ao obter/ler o Excel do plano anual (página inacessível, sem .xlsx, etc.)."""


# --- Normalização --------------------------------------------------------

def _clean(value) -> str:
    """Texto limpo (str vazia para None) — para os campos textuais do modelo."""
    if value is None:
        return ""
    return str(value).strip()


def _is_invitation(grant_type) -> bool:
    """True se a natureza do aviso é 'Convite'. Convites NÃO entram no plano (só concursos e
    outras naturezas abertas) — ver `sync_planned_grants`."""
    return "convite" in (grant_type or "").casefold()


def _parse_int(value) -> int | None:
    """Inteiro a partir da célula (tolera '12', 12, 12.0, ' 12 '); None se não for numérico."""
    if value in (None, ""):
        return None
    try:
        return int(float(str(value).strip().replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _parse_decimal(value) -> Decimal | None:
    """Converte um valor monetário do Excel em Decimal.

    Aceita números nativos do openpyxl e texto no formato português ('990 000 €',
    '1.234.567,89 €'). Remove símbolos/espaços, interpreta ',' como decimal e '.' como
    separador de milhares. None se vazio ou ilegível.
    """
    if value in (None, ""):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    cleaned_number = re.sub(r"[^\d,.\-]", "", str(value).replace("\xa0", " "))
    if not cleaned_number:
        return None
    if "," in cleaned_number and "." in cleaned_number:       # '1.234.567,89' → '.' milhares, ',' decimal
        cleaned_number = cleaned_number.replace(".", "").replace(",", ".")
    elif "," in cleaned_number:                  # '990000,50' → ',' decimal
        cleaned_number = cleaned_number.replace(",", ".")
    try:
        return Decimal(cleaned_number)
    except InvalidOperation:
        return None


def _parse_date(value) -> date | None:
    """Data a partir da célula: openpyxl devolve datetime nas células de data; caso contrário
    delega no parser tolerante partilhado (`common.dates.parse_date`)."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date(str(value))


# --- Obtenção do Excel ---------------------------------------------------

def _find_xlsx_url() -> str:
    """Descobre o URL do Excel: percorre todos os `<a>` da página e devolve o 1.º href que
    termina em '.xlsx' (resolvido para absoluto). Sem URL fixa."""
    try:
        http_response = requests.get(PLAN_URL, headers=_HEADERS, timeout=_TIMEOUT)
        http_response.raise_for_status()
    except requests.RequestException as exc:
        raise PlannedGrantsSyncError(f"Falha ao aceder ao plano anual: {exc}") from exc

    soup = BeautifulSoup(http_response.text, "html.parser")
    for anchor in soup.find_all("a", href=True):
        href = anchor["href"].strip()
        if href.lower().endswith(".xlsx"):
            return urljoin(PLAN_URL, href)
    raise PlannedGrantsSyncError("Nenhum ficheiro .xlsx encontrado na página do plano anual.")


def _download_workbook():
    """Descarrega o Excel para memória e devolve o workbook (só-leitura, valores calculados)."""
    url = _find_xlsx_url()
    try:
        http_response = requests.get(url, headers=_HEADERS, timeout=_TIMEOUT)
        http_response.raise_for_status()
    except requests.RequestException as exc:
        raise PlannedGrantsSyncError(f"Falha ao descarregar o Excel do plano anual: {exc}") from exc
    return load_workbook(io.BytesIO(http_response.content), read_only=True, data_only=True)


# --- Sincronização -------------------------------------------------------

def _row_to_fields(row) -> dict | None:
    """Mapeia uma linha (tuplo de valores das células) para {campo: valor} já normalizado.

    None se a linha não tiver um ID numérico (linhas vazias/rodapé são ignoradas)."""
    cells = list(row) if row is not None else []

    def cell(index: int):
        return cells[index] if index < len(cells) else None

    plan_id = _parse_int(cell(0))
    if plan_id is None:
        return None

    fields = {"plan_id": plan_id}
    for offset, name in enumerate(_COLUMN_ORDER, start=1):
        raw = cell(offset)
        if name in _DECIMAL_FIELDS:
            fields[name] = _parse_decimal(raw)
        elif name in _DATE_FIELDS:
            fields[name] = _parse_date(raw)
        else:
            fields[name] = _clean(raw)
    return fields


@transaction.atomic
def sync_planned_grants() -> None:
    """Sincroniza a tabela `PlannedGrant` com o Excel do plano anual.

    Para cada linha: cria se o `plan_id` não existir; se existir, atualiza APENAS os campos
    alterados; se estiver tudo igual, não escreve nada. Não devolve estatísticas.

    Os avisos de natureza 'Convite' são EXCLUÍDOS do plano: nunca são criados e, se um registo
    existente passar a convite numa sincronização, é removido (auto-correção).
    """
    workbook = _download_workbook()
    try:
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2, values_only=True):  # ignora o cabeçalho
            fields = _row_to_fields(row)
            if fields is None:
                continue
            plan_id = fields.pop("plan_id")
            if _is_invitation(fields["grant_type"]):
                # Nunca guardar convites — e remover um que porventura já lá esteja.
                PlannedGrant.objects.filter(plan_id=plan_id).delete()
                continue
            existing = PlannedGrant.objects.filter(plan_id=plan_id).first()
            if existing is None:
                PlannedGrant.objects.create(plan_id=plan_id, **fields)
                continue
            changed = [name for name, value in fields.items()
                       if getattr(existing, name) != value]
            if changed:
                for name in changed:
                    setattr(existing, name, fields[name])
                existing.save(update_fields=[*changed, "updated_at"])
    finally:
        workbook.close()


# --- Consulta ---------------------------------------------------------------

def upcoming_planned_grants(order_by: str | None = None):
    """Avisos previstos com abertura a partir de HOJE (esconde os já passados) — a base da
    listagem pública. Sem data de abertura conhecida também fica de fora: não há como provar
    que ainda está por vir. Fonte única partilhada com a newsletter (`coming_next_30_days`).

    `order_by` (ver `ORDERING`): abertura/fim previstos ou dotação; default (Meta.ordering) é
    abertura prevista mais cedo primeiro."""
    qs = PlannedGrant.objects.filter(expected_start__gte=date.today())
    if order_by in ORDERING:
        qs = qs.order_by(ORDERING[order_by])
    return qs


# --- Serialização --------------------------------------------------------

def serialize_planned_grant(pg: PlannedGrant) -> dict:
    """Aviso previsto em dict pronto para JSON (campos suficientes para a interface)."""
    return {
        "id": pg.id,
        "plan_id": pg.plan_id,
        "designation": pg.designation,
        "programme": pg.programme,
        "expected_start": pg.expected_start.isoformat() if pg.expected_start else None,
        "expected_end": pg.expected_end.isoformat() if pg.expected_end else None,
        "fund": pg.fund,
        "budget": float(pg.total_allocation) if pg.total_allocation is not None else None,
        "nuts": pg.nuts,
    }
